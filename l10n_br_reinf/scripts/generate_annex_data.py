#!/usr/bin/env python3
# Copyright 2026 KMEE - Luis Felipe Mileo <mileo@kmee.com.br>
# License AGPL-3 - See http://www.gnu.org/licenses/agpl-3.0.html
"""Generate the data files of the EFD-Reinf out of the official tables.

The source is the package "EFD-REINF Tabelas" published on the SPED portal, as
XLSX, and NOT the PDF of the manual: the manual only shows the natures that
generate a revenue code, its Annex I wraps the descriptions across lines in a
way no extraction survives, and it lags the tables by months.

Usage::

    python3 generate_annex_data.py <directory of the xlsx> <data directory>

It rewrites, in the data directory:

* ``l10n_br_reinf.nature.income.csv`` out of "Tabela 01", which is the complete
  list of natures of income, including the ones that suffer no withholding at
  all, such as 12001 Lucro e Dividendo;
* ``l10n_br_reinf.nature.income.tax.csv`` out of the tables of the R-4010,
  R-4020, R-4040 and R-4080, which map a nature to a tax, to a revenue code and
  to a period of collection, each line with its own validity.

Needs openpyxl, which is a dependency OF THIS SCRIPT and not of the module: the
module ships the generated CSV and never reads a spreadsheet.
"""

import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("This script needs openpyxl: pip install openpyxl")

# Tributo of the tables to the vocabulary of the module. The tables write the
# aggregated one in upper case, and the R-4040 and R-4080 do not break the
# withholding down by tax at all.
TAX = {
    "IRPF": "irpf",
    "IRPJ": "irpj",
    "RRA": "rra",
    "AGREGADO": "aggregated",
    "AGREGADO ": "aggregated",
    "CSLL": "csll",
    "COFINS": "cofins",
    "PIS/PASEP": "pis_pasep",
}

PERIODICITY = {
    "Mensal": "monthly",
    "Decendial": "ten_day",
    "Diário": "daily",
    "Diario": "daily",
    "Quinzenal": "fortnightly",
    "Semanal": "weekly",
}

# "Classificação tributária 85" of the table of the R-4020: whether the
# declarant is of the tax classification 85. It is what picks between the pair
# of revenue codes of a nature.
CLASSIFICATION = {"Sim": "yes", "Não": "no", "Nao": "no", "N/A": "na", None: ""}

# Which table belongs to which event, and where its columns are.
EVENT_TABLES = (
    # (file name, event, has tributo, has classification)
    ("EFD-REINF - Tabela R-4010.xlsx", "R-4010", True, False),
    ("EFD-REINF - Tabela R-4020.xlsx", "R-4020", True, True),
    ("EFD-REINF - Tabela R-4040.xlsx", "R-4040", False, False),
    ("EFD-REINF - Tabela R-4080.xlsx", "R-4080", False, False),
)


def iso_date(value):
    """The tables write dates as DD/MM/AAAA, and an open end as empty."""
    if not value:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if "/" in text:
        day, month, year = text.split("/")
        return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
    return text[:10]


def clean(value):
    return " ".join(str(value or "").split())


def read_natures(source):
    """Natures of income out of the Tabela 01."""
    path = source / "EFD-REINF - Tabela 01.xlsx"
    sheet = openpyxl.load_workbook(path, data_only=True).active
    natures = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        code = clean(row[1])
        if not code.isdigit() or len(code) != 5:
            continue
        # A nature that is reissued keeps the widest validity: the versioning
        # that matters for the declaration is the one of the mapping, which
        # carries the revenue code.
        name = clean(row[2])
        # Column "Tributo" of the Tabela 01: which taxes the nature admits. It
        # is the authorization of the aggregated withholding and it says which
        # components the aggregate carries, which is NOT always the three:
        # 15001, cooperatives of work, admits "IR, COFINS, PP, AGREGADO" with no
        # CSLL, because the art. 32 I of the Law 10.833 does not require it.
        admitted = clean(row[11]).upper()
        date_start, date_end = iso_date(row[12]), iso_date(row[13])
        if code in natures:
            previous = natures[code]
            natures[code] = (
                previous[0] or name,
                min(filter(None, [previous[1], date_start]), default=""),
                "" if not (previous[2] and date_end) else max(previous[2], date_end),
                previous[3] or admitted,
            )
            continue
        natures[code] = (name, date_start, date_end, admitted)
    return natures


def read_mappings(source):
    """Mappings of nature to tax, revenue code and period, per event."""
    mappings = []
    for file_name, event, has_tax, has_classification in EVENT_TABLES:
        sheet = openpyxl.load_workbook(source / file_name, data_only=True).active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            code = clean(row[0])
            if not code.isdigit() or len(code) != 5:
                continue
            index = 2
            foreign = ""
            if has_tax:
                foreign = clean(row[index])
                index += 1
                tax = TAX.get(clean(row[index]).upper(), "")
                index += 1
            else:
                tax = ""
            classification = ""
            if has_classification:
                classification = CLASSIFICATION.get(clean(row[index]), "")
                index += 1
            revenue_code = clean(row[index])
            index += 1
            periodicity = PERIODICITY.get(clean(row[index]), "")
            index += 1
            mappings.append(
                {
                    "nature": code,
                    "nature_name": clean(row[1]),
                    "event_type": event,
                    "tax_type": tax,
                    "foreign_taxation": "True" if foreign in ("Sim", "S") else "False",
                    "classification_indicator": classification,
                    "revenue_code": revenue_code,
                    "periodicity": periodicity,
                    "date_start": iso_date(row[index]),
                    "date_end": iso_date(row[index + 1]),
                }
            )
    return mappings


def write_natures(natures, data_dir):
    path = data_dir / "l10n_br_reinf.nature.income.csv"
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "id",
                "code",
                "name",
                "admitted_taxes",
                "date_start",
                "date_end",
            ],
            quoting=csv.QUOTE_ALL,
        )
        writer.writeheader()
        for code in sorted(natures):
            name, date_start, date_end, admitted = natures[code]
            writer.writerow(
                {
                    "id": f"nature_income_{code}",
                    "code": code,
                    "name": name,
                    "admitted_taxes": admitted,
                    "date_start": date_start,
                    "date_end": date_end,
                }
            )
    return path


def write_mappings(mappings, data_dir):
    path = data_dir / "l10n_br_reinf.nature.income.tax.csv"
    fields = [
        "id",
        "nature_income_id:id",
        "event_type",
        "tax_type",
        "foreign_taxation",
        "classification_indicator",
        "revenue_code",
        "periodicity",
        "date_start",
        "date_end",
    ]
    seen = {}
    rows = []
    for mapping in mappings:
        key = "_".join(
            [
                mapping["nature"],
                mapping["event_type"].replace("-", "").lower(),
                mapping["tax_type"] or "nd",
                "ext" if mapping["foreign_taxation"] == "True" else "nac",
                mapping["classification_indicator"] or "nd",
                mapping["revenue_code"] or "nd",
            ]
        )
        # A same key reissued with another validity gets a suffix, so the
        # external id stays stable and unique.
        seen[key] = seen.get(key, 0) + 1
        if seen[key] > 1:
            key = f"{key}_v{seen[key]}"
        row = {
            "id": f"nature_income_tax_{key}",
            "nature_income_id:id": f"l10n_br_reinf.nature_income_{mapping['nature']}",
        }
        row.update({field: mapping[field] for field in fields[2:]})
        rows.append(row)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        for row in sorted(rows, key=lambda item: item["id"]):
            writer.writerow(row)
    return path, rows


def recover_missing_natures(natures, mappings):
    """Natures used by a mapping table and absent from the Tabela 01.

    The published tables do disagree: 17547 is mapped in the table of the
    R-4020 and is not in the Tabela 01, which has 17546 and 17548. The mapping
    table carries the description of the nature in a column of its own, so the
    nature is recovered from there. That is the official text of a sibling
    table, not a description made up here, and the caller reports which ones
    had to be recovered.
    """
    recovered = []
    for mapping in mappings:
        code = mapping["nature"]
        if code in natures:
            continue
        natures[code] = (mapping["nature_name"], mapping["date_start"], "", "")
        recovered.append(code)
    return sorted(set(recovered))


def main():
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    source, data_dir = Path(sys.argv[1]), Path(sys.argv[2])
    natures = read_natures(source)
    mappings = read_mappings(source)
    missing = recover_missing_natures(natures, mappings)
    write_natures(natures, data_dir)
    _path, rows = write_mappings(mappings, data_dir)
    # stdout is the interface of a CLI script, and pylint_odoo refuses print.
    sys.stdout.write(f"natures: {len(natures)}\n")
    sys.stdout.write(f"mappings: {len(rows)}\n")
    if missing:
        sys.stdout.write(
            "natures recovered from the mapping tables because the Tabela 01 "
            f"does not list them: {missing}\n"
        )


if __name__ == "__main__":
    main()
